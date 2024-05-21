import logging
from flask import Flask, request, redirect, url_for, send_from_directory, render_template
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)

logging.basicConfig(level=logging.DEBUG)

# Configure o diretório de templates
app.template_folder = 'C://Desenvolvimento//comudFHub//Descktop//cadMem//spreadsheets//interface'

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.route('/spreadsheets/<path:filename>')
def send_spreadsheets(filename):
    return send_from_directory('spreadsheets', filename)

@app.route('/interface/<path:filename>')
def send_interface(filename):
    app.logger.debug(f'Trying to send file: {filename}')
    return send_from_directory('interface', filename)

@app.route('/submit', methods=['POST'])
def submit():
    nome = request.form['nome']
    idade = request.form['idade']
    cidade = request.form['cidade']
    
    # Caminho relativo para a planilha
    spreadsheet_path = 'spreadsheets//cadmem.xlsx'
    
    if os.path.exists(spreadsheet_path):
        wb = load_workbook(spreadsheet_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Pessoais"
        headers = ['Nome', 'Idade', 'Cidade']
        ws.append(headers)
    
    # Adicionar os dados do formulário na planilha
    ws.append([nome, idade, cidade])
    
    # Salvar a planilha
    wb.save(spreadsheet_path)
    
    # Redirecionar para a página HTML na pasta 'interface'
    return redirect(url_for('send_interface', filename='cad+mem.html'))

if __name__ == '__main__':
    app.run(debug=True)
